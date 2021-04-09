# -*- coding: UTF-8 -*-
# __CONFIG__ = "/etc/demo.json"

""" demo.py

Enjoy nice programming solutions for getting most powerfull tools for doing your tasks.
IP philosophy aims to define a simple and inovative way for making you efficient as never before:

    - enjoy coding in a clear, powerfull and simple way
    - enjoy learning with nice documentation and tutorials
    - enjoy progress and inovation working and exchanging with others
    - enjoy tools really designed for you and your concerns

:author: champion.nicolas@gmail.com
:licence: https://cecill.info/licences/Licence_CeCILL-C_V1-en.txt


Runtime directory
`````````````````
demo.py will run only within a specific directory given by demo.json
__CONFIG__ provide the path of this configuration file as a string
the current working directory will be set to the 'app' value given within demo.json
the following structure is hard coded and has to exists within:

    ./bottle_templates : root dir for getting bottle templates
    ./wsgi_scripts : 


Scripting small rules
``````````````````````
  ########################################################################################
 ##########  NEW SECTION  ###############################################################
########################################################################################

# highlighted topic :
# ```````````````````


Developpement topics
````````````````````
Incredible Progress defined its own Zen for coding:
    
    **THANKS FOR SAVING STABILITY, PERFORMANCE, CODE QUALITY**
    - Spend time and work hard in this purposes
    - Users concerns are the major topics, think first to cover it

    - Python is battery-included: first know it and use it
    - If not, think about a thin-pythonic-layer concept around
    - Keep code comprehensive, readable, maintenable by others anytime
    - Know what are coding agility best practices and apply it
    - All sophisticated things are bad
    - Others have to unsderstand your code, not you
    - Clean, fast, comprehensive are good
    - Provid both quick documentation in french and api documentation in english

    `Javascript`
    - Use JS for presentation purposes and html/css handlings matters
    - Keep a javascript-centric approche and code in a clean and comprehensive way
    - Try JS libraries freely but choose and test them carrefully for production
    - Use libraries provided within Ubuntu packages when available
    - Making elaborated JS code look for using TypeScript instead
    - Let middleware, data handlings, and calculations for python

TODO:
`````
    - implement cherrypy use instead of bottle
    - implement a CLI in good way
    - implement documentation tools
    - develop some nice excel examples with openpyxl

FIXME:
```````
    - bottle relics settings
"""

  ########################################################################################
 ##########  MODULES & FACILITIES #######################################################
########################################################################################

# import standard modules:
import os
from openpyxl import load_workbook
from datetime import datetime; from calendar import isleap

# import server/middleware facilities:
from sweet import _, webapp, html, Route
os.chdir(_["working_dir"])

sw.webapp.mount(
    sw.Route("/",sw.html("4i.generalquery.txt")),
)

sw.quickstart()
sw.exit()


  ########################################################################################
 ##########  BOTTLE RELICS  #############################################################
########################################################################################

# from cheroot import wsgi # cheroot is the CherryPy webserver usable with bottle
# from bottle import route, template, redirect, run, request, static_file, ServerAdapter


# # cheroot server implementation and interplay with bottle 12x:
# # based on https://groups.google.com/forum/#!topic/bottlepy/rffluU8aOMI
# class CherootServer(ServerAdapter):
#     def run(self, handler):
#         server = wsgi.Server((self.host, self.port), handler)
#         try:
#             server.start()
#         finally:
#             server.stop()


# # configuration
# with open(__CONFIG__) as file:
#     config = json.load(file)
# try:

#     server_bottle = { "server": CherootServer,
#         "host": config['bottle']['host'],
#         "port": config['bottle']['port'],
#         "debug": True,
#         "reloader": False } # NOTE: for dev purposes only
    
#     static_bottle = config['bottle']['static_bottle'] #type: dict

# except:
#     raise KeyError("loading %s configuration file" % __CONFIG__)


# # provide convenient function for rendering templates or urls:
# def html(source, *args, **kwargs): 
#     if "://" in source:
#         print("[DEMO] redirect to", source)
#         return redirect(source)
#     path = os.path.join('bottle_templates', source)
#     print("[DEMO] load bottle template:", path)
#     return template(path, *args, **kwargs)


# # Route urls for static files:
# @route('/<filepath:path>')
# def server_static(filepath):
#     root = "" #default

#     # check for a rule from config:
#     # autoset usual resources for 'simplifying html' purpose
#     for pattern in static_bottle:
#         if pattern == filepath  or pattern in filepath:
#             root = static_bottle[pattern]
#             break

#     # load static file:
#     print("[DEMO] load static file:", os.path.join(root, filepath)) 
#     return static_file(filepath, root=root)


  ########################################################################################
 ##########  DATA HANDLING  #############################################################
########################################################################################

class TimetableIO():

    def __init__(self):

        self.database = sweet.database
        self.user = "Julie Latour"
        self.year = 2020
        self.team = [ "Julie Latour",
            "Marcel Dupont",
            "David Martin",
            "Marie Aimé",
            "Sophie Duprès" ]

    def html(self):

        self.timetable_dict = {}
        self.months_len = [31,29 if isleap(self.year) else 28,31,30,31,30,31,31,30,31,30,31]

        for month_nbr, month_len in enumerate(self.months_len) :
            for day_nbr in range(month_len) :

                date_id = '-'.join((str(day_nbr+1), str(month_nbr+1), str(self.year)))
                try: dbitem = self.database["timetable"].find_one(
                    {"user":self.user,"date":date_id})['status']
                except: dbitem = ''
                self.timetable_dict[date_id] = dbitem

        print("[DEMO] fetch timetable for  user:", self.user, " year:", self.year) 
        
        return sweet.html("timetable.txt",
            
            timetable= self.timetable_dict,
            user= self.user,
            year= self.year,
            months_len= self.months_len,
            datetime= datetime )

    def update(self, date_id, status):

        # insert or update data within MongoDB:
        query = { "date": date_id, "user": self.user }
        dbitem = self.database["timetable"].find_one(query, { "_id": 0 })

        if dbitem is None :
            print('[DEMO] insert  id=',date_id," find_one=",dbitem," new=",status)
            self.database["timetable"].insert_one({
                "date": date_id,
                "user": self.user,
                "status": status })

        elif dbitem['status'] != status :
            print('[DEMO] update  id=',date_id," find_one=",dbitem," new=",status)
            newval = { "$set": { "status": status } }
            self.database["timetable"].update_one(query, newval)

    def change(self, user, year):

        self.user = user
        self.year = int(year[-4:])
        return self.html()
     
    def excel(self):

        wb = load_workbook(filename="/opt/incredible/excel/xlsx_templates/timetable.xlsx")
        ws = wb.worksheets

        for month in range(1): # not 12 for test
            ws[month]["AH4"].value = self.year

            for index, user in enumerate(self.team):
                row = 7+index
                ws[month].cell(row,2).value = user
                
                for day in range(31):
                    date_id = '-'.join((str(day+1), str(month+1), str(self.year)))
                    dbitem = self.database["timetable"].find_one({ "user": user, "date": date_id })
                    if not dbitem: continue
                    else: dbitem = dbitem["status"]

                    col = 3+day
                    ws[month].cell(row,col).value = dbitem

        wb.save(filename="/opt/incredible/excel/.temporary_files/timetable_v1.xlsx")
        
        os.system( "".join(('cmd.exe /c start excel.exe "', rootfs,
            '/opt/incredible/excel/.temporary_files/timetable_v1.xlsx"')) )


  ########################################################################################
 ##########  URLS ROUTING  ##############################################################
########################################################################################

# class HtmlTimetable():

#     io = TimetableIO()

#     @sweetheart.expose()
#     def index(self):
#         return self.io.html()

#     @sweetheart.expose()
#     def excel(self):
#         self.io.excel()

#     @sweetheart.expose()
#     def test(self, date_id, status):
#         sweetheart.echo("date_id:", date_id, "status", status)


  ########################################################################################
 ##########  COMMAND LINE  ##############################################################
########################################################################################

if __name__ == "__main__":

        io = TimetableIO()

        def timetable(request):
            return io.html()
        
        def tt_excel(request):
            return io.excel()

        def tt_test(request):
            sweet.echo("test")

        sweet.quickstart({
            "/": sweet.welcome,
            "/timetable": timetable,
            "/timetable/excel": tt_excel,
            "/timetable_test": tt_test })
        

        # sweetheart.mount({
        #     "/": sweetheart.HtmlRoot(),
        #     "/timetable": HtmlTimetable() })

        # sweetheart.dispatch({
        #     "/timetable/exchange": { 
        #     "GET": HtmlTimetable.io.change,
        #     "PUT": HtmlTimetable.io.update } })

        # sweetheart.start()


  ########################################################################################
 ##########  FOR TEST  ##################################################################
########################################################################################

# import requests #> the http protocol for human beings
# from bs4 import BeautifulSoup #> for handling html documents in convenient way


# provide tools for building documentation
# def record_media(
#         id:str, 
#         path:str, 
#         description:str="", 
#         tags:str="" ):

#     with open(path,'rb') as file:
#         echo("record file type:", type(file))

#         if database["media"].find_one({"_id": id}):

#             database["media"].update(
#                 {"_id": id},
#                 { "$set": { 
#                     "document": file }} )
#         else:
#             database["media"].insert_one({
#                 "_id": id,
#                 "tags": tags,
#                 "description": description,
#                 "document":  file.read() })


# @cherrypy.expose
    # def learn_python(self):
    #     """NOTE: for test only"""
    #     return html({
    #         "do": "DOC_IMPORT",
    #         "url": "https://python.doctor/page-apprendre-listes-list-tableaux-tableaux-liste-array-python-cours-debutant",
    #         "find": ("div", {"id":"page"}),
    #     })


# elif isinstance(source, dict):
#     #NOTE: for test only
#     if source.get("do") == "DOC_IMPORT":

#         html_doc = requests.get(source["url"])
#         find_args = source["find"]
#         soup = BeautifulSoup(html_doc.text, "html.parser")

#         return template(
#             "bottle_templates/document.txt",
#             text= soup.find(*find_args),
#             default_libs= _config_["webapp"]["default_libs"],
#             **kwargs )
